#!/usr/bin/env python3
"""Generate test Excel files for product import testing"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Columns
COLUMNS = [
    'name', 'price', 'sku', 'category_slug', 'supplier_name',
    'description', 'images', 'in_stock', 'min_order', 'specifications', 'tags'
]

def style_header(ws):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, column in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 45  # name
    ws.column_dimensions['B'].width = 12  # price
    ws.column_dimensions['C'].width = 15  # sku
    ws.column_dimensions['D'].width = 18  # category_slug
    ws.column_dimensions['E'].width = 20  # supplier_name
    ws.column_dimensions['F'].width = 60  # description
    ws.column_dimensions['G'].width = 50  # images
    ws.column_dimensions['H'].width = 10  # in_stock
    ws.column_dimensions['I'].width = 12  # min_order
    ws.column_dimensions['J'].width = 50  # specifications
    ws.column_dimensions['K'].width = 30  # tags


# ========== FILE 1: PERFECT PRODUCTS ==========
perfect_products = [
    {
        'name': 'Смартфон Xiaomi Redmi Note 13 Pro 8/256GB',
        'price': 28990,
        'sku': 'PHONE-001',
        'category_slug': 'smartphones',
        'supplier_name': 'Xiaomi Official',
        'description': 'Флагманский смартфон с камерой 200 МП и AMOLED-дисплеем 120 Гц. Мощный процессор Snapdragon обеспечивает плавную работу приложений. Быстрая зарядка 67 Вт полностью заряжает устройство за 45 минут.',
        'images': 'https://images.unsplash.com/photo-1511707171634-5f897ff02aa9?w=800,https://images.unsplash.com/photo-1592899677977-9c10ca588bbd?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Экран:6.67" AMOLED 120Hz|Память:8/256 GB|Камера:200+8+2 МП|Батарея:5000 мАч|Процессор:Snapdragon 7s Gen 2',
        'tags': 'смартфон,xiaomi,redmi,android,5g,быстрая зарядка'
    },
    {
        'name': 'Беспроводные наушники Apple AirPods Pro 2',
        'price': 24990,
        'sku': 'AUDIO-001',
        'category_slug': 'headphones',
        'supplier_name': 'Apple Russia',
        'description': 'Премиальные TWS наушники с активным шумоподавлением и пространственным звуком. Адаптивный режим прозрачности позволяет слышать окружение. До 6 часов воспроизведения, 30 часов с кейсом.',
        'images': 'https://images.unsplash.com/photo-1606220945770-b5b6c2c55bf1?w=800,https://images.unsplash.com/photo-1600294037681-c80b4cb5b434?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Тип:TWS вкладыши|ANC:Да|Bluetooth:5.3|Время работы:6+24 ч|Защита:IPX4|Чип:Apple H2',
        'tags': 'наушники,apple,airpods,tws,anc,беспроводные'
    },
    {
        'name': 'Умные часы Samsung Galaxy Watch 6 Classic 47mm',
        'price': 34990,
        'sku': 'WATCH-001',
        'category_slug': 'electronics',
        'supplier_name': 'Samsung Electronics',
        'description': 'Премиальные смарт-часы с вращающимся безелем и Super AMOLED экраном. Расширенный мониторинг здоровья: ЭКГ, давление, сон, стресс. Автономность до 40 часов в обычном режиме.',
        'images': 'https://images.unsplash.com/photo-1546868871-7041f2a55e12?w=800,https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Дисплей:1.47" AMOLED|Процессор:Exynos W930|Память:2/16 GB|Батарея:425 мАч|Защита:5ATM+IP68|ОС:Wear OS 4',
        'tags': 'часы,samsung,galaxy watch,smartwatch,wear os,здоровье'
    },
    {
        'name': 'Робот-пылесос Roborock S8 Pro Ultra',
        'price': 89990,
        'sku': 'HOME-001',
        'category_slug': 'home-appliances',
        'supplier_name': 'Roborock Tech',
        'description': 'Топовый робот-пылесос с док-станцией самоочистки и автоматической мойкой тряпок. LiDAR навигация строит точные карты помещений. Мощность всасывания 6000 Па справится с любым мусором.',
        'images': 'https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Мощность:6000 Па|Навигация:LiDAR 3D|Батарея:5200 мАч|Пылесборник:400 мл|Бак для воды:300 мл|Площадь:300 м²',
        'tags': 'робот-пылесос,roborock,умный дом,уборка,моющий'
    },
    {
        'name': 'Автоматическая кофемашина De\'Longhi Magnifica S',
        'price': 42990,
        'sku': 'KITCHEN-001',
        'category_slug': 'kitchen',
        'supplier_name': 'De\'Longhi Russia',
        'description': 'Полностью автоматическая кофемашина с керамическими жерновами. Готовит эспрессо, капучино и латте одним нажатием. Встроенный капучинатор создает идеальную молочную пену.',
        'images': 'https://images.unsplash.com/photo-1517668808822-9ebb02f2a0e6?w=800,https://images.unsplash.com/photo-1495474472287-4d71bcdd2085?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Давление:15 бар|Мощность:1450 Вт|Ёмкость воды:1.8 л|Кофемолка:Керамическая|Напитки:Эспрессо,капучино,латте|Управление:Сенсорное',
        'tags': 'кофемашина,delonghi,кофе,эспрессо,капучино,кухня'
    },
    {
        'name': 'Фен-стайлер Dyson Airwrap Complete Long',
        'price': 54990,
        'sku': 'BEAUTY-001',
        'category_slug': 'beauty',
        'supplier_name': 'Dyson Official',
        'description': 'Революционный стайлер для создания локонов, волн и объема без экстремального нагрева. Технология Coanda притягивает волосы к насадке. В комплекте 6 насадок для разных укладок.',
        'images': 'https://images.unsplash.com/photo-1522338140262-f46f5913618a?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Мощность:1300 Вт|Насадки:6 шт|Режимы:3 температуры|Длина шнура:2.7 м|Технология:Coanda Air|Для волос:От 20 см',
        'tags': 'фен,стайлер,dyson,airwrap,укладка,красота'
    },
    {
        'name': 'Фитнес-браслет Xiaomi Mi Band 8 Pro',
        'price': 4990,
        'sku': 'FITNESS-001',
        'category_slug': 'sports',
        'supplier_name': 'Xiaomi Official',
        'description': 'Продвинутый фитнес-трекер с AMOLED экраном 1.74" и встроенным GPS. Отслеживает 150+ режимов тренировок, пульс, SpO2 и качество сна. Водонепроницаемость 5ATM для плавания.',
        'images': 'https://images.unsplash.com/photo-1575311373937-040b8e1fd5b6?w=800,https://images.unsplash.com/photo-1576243345690-4e4b79b63288?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Дисплей:1.74" AMOLED|GPS:Встроенный|Батарея:14 дней|Защита:5ATM|Датчики:Пульс,SpO2,акселерометр|Тренировки:150+',
        'tags': 'фитнес-браслет,xiaomi,mi band,трекер,спорт,здоровье'
    },
    {
        'name': 'Портативная колонка JBL Charge 5',
        'price': 15990,
        'sku': 'AUDIO-002',
        'category_slug': 'electronics',
        'supplier_name': 'JBL Harman',
        'description': 'Мощная портативная колонка с глубоким басом и защитой IP67 от воды и пыли. Powerbank функция для зарядки смартфона. До 20 часов воспроизведения музыки без подзарядки.',
        'images': 'https://images.unsplash.com/photo-1608043152269-423dbba4e7e1?w=800,https://images.unsplash.com/photo-1589003077984-894e133dabab?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Мощность:40 Вт|Bluetooth:5.1|Батарея:20 ч|Защита:IP67|Функции:PartyBoost,Powerbank|Вес:960 г',
        'tags': 'колонка,jbl,bluetooth,портативная,музыка,водонепроницаемая'
    }
]


# ========== FILE 2: ERROR PRODUCTS ==========
error_products = [
    # 1. Нормальный товар для контраста
    {
        'name': 'Наушники Sony WH-1000XM5',
        'price': 34990,
        'sku': 'TEST-001',
        'category_slug': 'headphones',
        'supplier_name': 'Sony Russia',
        'description': 'Флагманские наушники с лучшим шумоподавлением в классе.',
        'images': 'https://images.unsplash.com/photo-1505740420928-5e560c06d30e?w=800',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': 'Тип:Накладные|ANC:Да|Батарея:30 ч',
        'tags': 'наушники,sony,anc'
    },
    # 2. Пустое название
    {
        'name': '',  # ERROR: Empty name
        'price': 5990,
        'sku': 'TEST-002',
        'category_slug': 'electronics',
        'supplier_name': 'Test Supplier',
        'description': 'Товар без названия для тестирования валидации.',
        'images': '',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': '',
        'tags': ''
    },
    # 3. Некорректная цена (текст)
    {
        'name': 'Товар с текстовой ценой',
        'price': 'бесплатно',  # ERROR: Text instead of number
        'sku': 'TEST-003',
        'category_slug': 'electronics',
        'supplier_name': 'Test Supplier',
        'description': 'Товар с некорректной ценой для тестирования.',
        'images': '',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': '',
        'tags': ''
    },
    # 4. Отрицательная цена
    {
        'name': 'Товар с отрицательной ценой',
        'price': -500,  # ERROR: Negative price
        'sku': 'TEST-004',
        'category_slug': 'electronics',
        'supplier_name': 'Test Supplier',
        'description': 'Товар с отрицательной ценой для тестирования.',
        'images': '',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': '',
        'tags': ''
    },
    # 5. Дубликат SKU
    {
        'name': 'Товар с дубликатом SKU',
        'price': 9990,
        'sku': 'TEST-001',  # ERROR: Duplicate SKU (same as row 1)
        'category_slug': 'electronics',
        'supplier_name': 'Test Supplier',
        'description': 'Товар с дублирующимся артикулом.',
        'images': '',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': '',
        'tags': ''
    },
    # 6. Несуществующая категория
    {
        'name': 'Товар с несуществующей категорией',
        'price': 7990,
        'sku': 'TEST-006',
        'category_slug': 'unknown-category',  # WARNING: Non-existent category
        'supplier_name': 'Test Supplier',
        'description': 'Товар с категорией, которой нет в базе.',
        'images': '',
        'in_stock': 'true',
        'min_order': 1,
        'specifications': '',
        'tags': ''
    }
]


def create_xlsx(filename, products):
    """Create Excel file with products"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Add headers
    style_header(ws)

    # Add data
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product['name'])
        ws.cell(row=row_num, column=2, value=product['price'])
        ws.cell(row=row_num, column=3, value=product['sku'])
        ws.cell(row=row_num, column=4, value=product['category_slug'])
        ws.cell(row=row_num, column=5, value=product['supplier_name'])
        ws.cell(row=row_num, column=6, value=product['description'])
        ws.cell(row=row_num, column=7, value=product['images'])
        ws.cell(row=row_num, column=8, value=product['in_stock'])
        ws.cell(row=row_num, column=9, value=product['min_order'])
        ws.cell(row=row_num, column=10, value=product['specifications'])
        ws.cell(row=row_num, column=11, value=product['tags'])

        # Wrap text for description
        ws.cell(row=row_num, column=6).alignment = Alignment(wrap_text=True)

    wb.save(filename)
    print(f"✅ Created: {filename}")


if __name__ == '__main__':
    create_xlsx('test_products_perfect.xlsx', perfect_products)
    create_xlsx('test_products_errors.xlsx', error_products)
    print("\n🎉 Both test files created successfully!")
